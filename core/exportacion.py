"""
Exportación de calificaciones a Excel.

Cada exportación genera un libro con dos hojas:
  - "Calificaciones": alumnos en filas, criterios en columnas (con su peso
    en la materia junto al código) y la nota final de la materia en la
    última columna. Cada celda combina nota numérica y calificación
    cualitativa (ej. "7,5 (NT)"), con el color degradado rojo-verde como
    relleno de la celda — igual que se ve en pantalla.
  - "Traspuesta": los mismos datos con filas y columnas intercambiadas
    (criterios en filas, alumnos en columnas), como la hoja TRASPUESTA
    del Excel original, para entregar a Jefatura de Estudios.

Se usa tanto para una evaluación normal (1EVA/2EVA/3EVA) como para FINAL.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.calificacion import calificacion_cualitativa, color_hex_nota
from core.database import BaseDatosCurso, Criterio, Evaluacion, Materia

FUENTE_CABECERA = Font(bold=True, color="FFFFFF")
RELLENO_CABECERA = PatternFill("solid", start_color="1D4E89")
FUENTE_NOTA_FINAL = Font(bold=True)
RELLENO_SIN_NOTA = PatternFill("solid", start_color="F2F2F2")
ALINEACION_CENTRO = Alignment(horizontal="center", vertical="center")


def _texto_y_relleno_criterio(valor: float | None) -> tuple[str, PatternFill | None]:
    """Para celdas de CRITERIO: solo el número con dos decimales (coma
    española) y el color degradado de fondo. Cadena vacía si no hay nota.
    """
    if valor is None:
        return "", RELLENO_SIN_NOTA
    texto = f"{valor:.2f}".replace(".", ",")
    color_hex = color_hex_nota(valor)
    relleno = PatternFill("solid", start_color=color_hex) if color_hex else None
    return texto, relleno


def _textos_y_relleno_nota_final(
    valor: float | None,
) -> tuple[str, str, PatternFill | None]:
    """Para la NOTA FINAL: devuelve (texto_numero, texto_letra, relleno),
    pensados para dos celdas separadas.
    """
    if valor is None:
        return "", "", RELLENO_SIN_NOTA
    texto_numero = f"{valor:.2f}".replace(".", ",")
    texto_letra = calificacion_cualitativa(valor)
    color_hex = color_hex_nota(valor)
    relleno = PatternFill("solid", start_color=color_hex) if color_hex else None
    return texto_numero, texto_letra, relleno


def _nombres_alumnos_y_notas(
    base_datos: BaseDatosCurso,
    materia: Materia,
    criterios: list[Criterio],
    incluir_solo_evaluables: Evaluacion | None,
) -> tuple[list, dict, dict]:
    """Reúne la lista de alumnos a exportar junto con sus notas de criterio
    y de materia, usando las funciones de cálculo que correspondan según
    si se exporta una evaluación normal o FINAL.

    Si incluir_solo_evaluables es una Evaluacion (1EVA/2EVA/3EVA), se
    excluyen los alumnos que todavía no estaban dados de alta en ella.
    Si es None, se exporta FINAL (incluye a todo el alumnado siempre).
    """
    if incluir_solo_evaluables is not None:
        vista = base_datos.listar_alumnos_para_evaluacion(materia.id, incluir_solo_evaluables.nombre)
        alumnos = [alumno for alumno, evaluable in vista if evaluable]
        notas_criterio = base_datos.calcular_notas_criterios_evaluacion(
            incluir_solo_evaluables.id, materia.id
        )
        notas_materia = base_datos.calcular_notas_materia_evaluacion(
            incluir_solo_evaluables.id, materia.id
        )
    else:
        alumnos = base_datos.listar_alumnos(materia.id)
        notas_criterio = base_datos.calcular_notas_criterios_final(materia.id)
        notas_materia = base_datos.calcular_notas_materia_final(materia.id)

    return alumnos, notas_criterio, notas_materia


def _construir_libro(
    base_datos: BaseDatosCurso,
    materia: Materia,
    incluir_solo_evaluables: Evaluacion | None,
) -> Workbook:
    criterios = base_datos.listar_criterios(materia.id)
    alumnos, notas_criterio, notas_materia = _nombres_alumnos_y_notas(
        base_datos, materia, criterios, incluir_solo_evaluables
    )

    libro = Workbook()
    hoja_normal = libro.active
    hoja_normal.title = "Calificaciones"
    hoja_traspuesta = libro.create_sheet("Traspuesta")

    # -- hoja normal: alumnos en filas, criterios en columnas --
    # Fila 1: cabecera de texto. Fila 2: código del criterio (1.1, 1.2...).
    # Fila 3: peso del criterio (solo el número). A partir de la fila 4,
    # un alumno por fila.
    encabezados = (
        ["Apellidos", "Nombre"] + [""] * len(criterios) + ["NOTA FINAL", "CALIFICACIÓN"]
    )
    col_final_numero = 3 + len(criterios)
    col_final_letra = col_final_numero + 1

    for columna, titulo in enumerate(encabezados, start=1):
        celda = hoja_normal.cell(row=1, column=columna, value=titulo)
        celda.font = FUENTE_CABECERA
        celda.fill = RELLENO_CABECERA
        celda.alignment = ALINEACION_CENTRO
    hoja_normal.cell(row=1, column=1, value="Apellidos")
    hoja_normal.cell(row=1, column=2, value="Nombre")

    fila_codigo = 2
    fila_peso = 3
    hoja_normal.cell(row=fila_codigo, column=1, value="Criterio")
    hoja_normal.cell(row=fila_peso, column=1, value="Peso")
    for fila_etiqueta in (fila_codigo, fila_peso):
        celda_et = hoja_normal.cell(row=fila_etiqueta, column=1)
        celda_et.font = FUENTE_NOTA_FINAL
    for col_indice, criterio in enumerate(criterios, start=3):
        celda_codigo = hoja_normal.cell(row=fila_codigo, column=col_indice, value=criterio.codigo)
        celda_codigo.font = FUENTE_NOTA_FINAL
        celda_codigo.alignment = ALINEACION_CENTRO
        celda_peso = hoja_normal.cell(row=fila_peso, column=col_indice, value=criterio.peso)
        celda_peso.alignment = ALINEACION_CENTRO

    fila_primer_alumno = 4
    for indice_alumno, alumno in enumerate(alumnos):
        fila_indice = fila_primer_alumno + indice_alumno
        hoja_normal.cell(row=fila_indice, column=1, value=alumno.apellidos)
        hoja_normal.cell(row=fila_indice, column=2, value=alumno.nombre)
        for col_indice, criterio in enumerate(criterios, start=3):
            valor = notas_criterio.get((criterio.id, alumno.id))
            texto, relleno = _texto_y_relleno_criterio(valor)
            celda = hoja_normal.cell(row=fila_indice, column=col_indice, value=texto)
            celda.alignment = ALINEACION_CENTRO
            if relleno is not None:
                celda.fill = relleno

        valor_final = notas_materia.get(alumno.id)
        texto_numero, texto_letra, relleno_final = _textos_y_relleno_nota_final(valor_final)
        celda_numero = hoja_normal.cell(row=fila_indice, column=col_final_numero, value=texto_numero)
        celda_letra = hoja_normal.cell(row=fila_indice, column=col_final_letra, value=texto_letra)
        for celda_f in (celda_numero, celda_letra):
            celda_f.font = FUENTE_NOTA_FINAL
            celda_f.alignment = ALINEACION_CENTRO
            if relleno_final is not None:
                celda_f.fill = relleno_final

    hoja_normal.column_dimensions["A"].width = 22
    hoja_normal.column_dimensions["B"].width = 18
    for col_indice in range(3, col_final_letra + 1):
        hoja_normal.column_dimensions[get_column_letter(col_indice)].width = 16
    hoja_normal.freeze_panes = f"C{fila_primer_alumno}"

    # -- hoja traspuesta: criterios en filas, alumnos en columnas --
    # Columna A: código del criterio. Columna B: peso. A partir de la
    # columna C, un alumno por columna.
    nombres_alumnos = [f"{a.apellidos}, {a.nombre}".strip(", ") for a in alumnos]
    encabezados_t = ["Criterio", "Peso"] + nombres_alumnos
    for columna, titulo in enumerate(encabezados_t, start=1):
        celda = hoja_traspuesta.cell(row=1, column=columna, value=titulo)
        celda.font = FUENTE_CABECERA
        celda.fill = RELLENO_CABECERA
        celda.alignment = ALINEACION_CENTRO

    col_primer_alumno = 3

    for fila_indice, criterio in enumerate(criterios, start=2):
        hoja_traspuesta.cell(row=fila_indice, column=1, value=criterio.codigo)
        hoja_traspuesta.cell(row=fila_indice, column=2, value=criterio.peso)
        for indice_alumno, alumno in enumerate(alumnos):
            col_indice = col_primer_alumno + indice_alumno
            valor = notas_criterio.get((criterio.id, alumno.id))
            texto, relleno = _texto_y_relleno_criterio(valor)
            celda = hoja_traspuesta.cell(row=fila_indice, column=col_indice, value=texto)
            celda.alignment = ALINEACION_CENTRO
            if relleno is not None:
                celda.fill = relleno

    fila_nota_final = 2 + len(criterios)
    fila_calificacion = fila_nota_final + 1
    celda_etiqueta_numero = hoja_traspuesta.cell(row=fila_nota_final, column=1, value="NOTA FINAL")
    celda_etiqueta_letra = hoja_traspuesta.cell(row=fila_calificacion, column=1, value="CALIFICACIÓN")
    celda_etiqueta_numero.font = FUENTE_NOTA_FINAL
    celda_etiqueta_letra.font = FUENTE_NOTA_FINAL
    for indice_alumno, alumno in enumerate(alumnos):
        col_indice = col_primer_alumno + indice_alumno
        valor_final = notas_materia.get(alumno.id)
        texto_numero, texto_letra, relleno_final = _textos_y_relleno_nota_final(valor_final)
        celda_numero = hoja_traspuesta.cell(row=fila_nota_final, column=col_indice, value=texto_numero)
        celda_letra = hoja_traspuesta.cell(row=fila_calificacion, column=col_indice, value=texto_letra)
        for celda_f in (celda_numero, celda_letra):
            celda_f.font = FUENTE_NOTA_FINAL
            celda_f.alignment = ALINEACION_CENTRO
            if relleno_final is not None:
                celda_f.fill = relleno_final

    hoja_traspuesta.column_dimensions["A"].width = 14
    hoja_traspuesta.column_dimensions["B"].width = 10
    for indice_alumno in range(len(alumnos)):
        col_letra = get_column_letter(col_primer_alumno + indice_alumno)
        hoja_traspuesta.column_dimensions[col_letra].width = 18
    hoja_traspuesta.freeze_panes = f"{get_column_letter(col_primer_alumno)}2"

    return libro


def _anadir_hoja_trazabilidad_evaluacion(libro: Workbook, base_datos: BaseDatosCurso, materia: Materia, evaluacion: Evaluacion):
    """Añade al libro una hoja "Trazabilidad" que muestra, para cada
    criterio (filas) y cada instrumento de evaluación de esta evaluación
    (columnas), si ese instrumento evalúa ese criterio (y con qué peso).
    """
    criterios, instrumentos, pesos = base_datos.trazabilidad_criterios_instrumentos(evaluacion.id, materia.id)
    hoja = libro.create_sheet("Trazabilidad")

    encabezados = ["Criterio"] + [instrumento.nombre for instrumento in instrumentos]
    for columna, titulo in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=columna, value=titulo)
        celda.font = FUENTE_CABECERA
        celda.fill = RELLENO_CABECERA
        celda.alignment = ALINEACION_CENTRO

    relleno_si = PatternFill("solid", start_color="C9EAE0")
    relleno_no = PatternFill("solid", start_color="F2F2F2")

    for fila_indice, criterio in enumerate(criterios, start=2):
        celda_criterio = hoja.cell(row=fila_indice, column=1, value=criterio.codigo)
        celda_criterio.font = FUENTE_NOTA_FINAL
        celda_criterio.alignment = ALINEACION_CENTRO
        for col_indice, instrumento in enumerate(instrumentos, start=2):
            peso = pesos.get((criterio.id, instrumento.id))
            if peso is None:
                celda = hoja.cell(row=fila_indice, column=col_indice, value="No")
                celda.fill = relleno_no
            else:
                celda = hoja.cell(row=fila_indice, column=col_indice, value=f"Sí ({peso:g})")
                celda.fill = relleno_si
            celda.alignment = ALINEACION_CENTRO

    hoja.column_dimensions["A"].width = 14
    for col_indice in range(2, len(instrumentos) + 2):
        hoja.column_dimensions[get_column_letter(col_indice)].width = 18
    hoja.freeze_panes = "B2"


def _anadir_hoja_trazabilidad_final(libro: Workbook, base_datos: BaseDatosCurso, materia: Materia):
    """Añade al libro una hoja "Trazabilidad" para FINAL: para cada
    criterio (filas) y cada evaluación 1EVA/2EVA/3EVA (columnas), indica
    si ese criterio tuvo alguna nota calculada en esa evaluación.
    """
    criterios, evaluaciones, evaluado = base_datos.trazabilidad_criterios_evaluaciones(materia.id)
    hoja = libro.create_sheet("Trazabilidad")

    encabezados = ["Criterio"] + [ev.nombre for ev in evaluaciones] + ["Nº de evaluaciones"]
    for columna, titulo in enumerate(encabezados, start=1):
        celda = hoja.cell(row=1, column=columna, value=titulo)
        celda.font = FUENTE_CABECERA
        celda.fill = RELLENO_CABECERA
        celda.alignment = ALINEACION_CENTRO

    relleno_si = PatternFill("solid", start_color="C9EAE0")
    relleno_no = PatternFill("solid", start_color="F2F2F2")

    for fila_indice, criterio in enumerate(criterios, start=2):
        celda_criterio = hoja.cell(row=fila_indice, column=1, value=criterio.codigo)
        celda_criterio.font = FUENTE_NOTA_FINAL
        celda_criterio.alignment = ALINEACION_CENTRO

        contador_evaluaciones = 0
        for col_indice, evaluacion in enumerate(evaluaciones, start=2):
            evaluado_aqui = evaluado.get((criterio.id, evaluacion.id), False)
            if evaluado_aqui:
                contador_evaluaciones += 1
            celda = hoja.cell(row=fila_indice, column=col_indice, value="Sí" if evaluado_aqui else "No")
            celda.fill = relleno_si if evaluado_aqui else relleno_no
            celda.alignment = ALINEACION_CENTRO

        col_contador = len(evaluaciones) + 2
        celda_contador = hoja.cell(row=fila_indice, column=col_contador, value=contador_evaluaciones)
        celda_contador.font = FUENTE_NOTA_FINAL
        celda_contador.alignment = ALINEACION_CENTRO

    hoja.column_dimensions["A"].width = 14
    for col_indice in range(2, len(evaluaciones) + 3):
        hoja.column_dimensions[get_column_letter(col_indice)].width = 18
    hoja.freeze_panes = "B2"


def exportar_evaluacion(
    base_datos: BaseDatosCurso, materia: Materia, evaluacion: Evaluacion, ruta_destino: str | Path
) -> Path:
    """Exporta las calificaciones de una evaluación normal (1EVA/2EVA/3EVA),
    junto con una hoja de trazabilidad (qué instrumento evalúa cada criterio).
    """
    libro = _construir_libro(base_datos, materia, incluir_solo_evaluables=evaluacion)
    _anadir_hoja_trazabilidad_evaluacion(libro, base_datos, materia, evaluacion)
    ruta_destino = Path(ruta_destino)
    libro.save(ruta_destino)
    return ruta_destino


def exportar_final(base_datos: BaseDatosCurso, materia: Materia, ruta_destino: str | Path) -> Path:
    """Exporta las calificaciones de la evaluación FINAL (todo el curso),
    junto con una hoja de trazabilidad (en qué evaluación se evaluó cada
    criterio: 1EVA, 2EVA, 3EVA, o varias de ellas).
    """
    libro = _construir_libro(base_datos, materia, incluir_solo_evaluables=None)
    _anadir_hoja_trazabilidad_final(libro, base_datos, materia)
    ruta_destino = Path(ruta_destino)
    libro.save(ruta_destino)
    return ruta_destino
